import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import xml.etree.ElementTree as ElementTree
from openpyxl import Workbook


def process_file(filename):
    try:
        wb = Workbook()
        ws = wb.active
        tree = ElementTree.parse(filename)
        ns = {'ns': 'http://www.demandware.com/xml/impex/catalog/2006-10-31'}
        products = tree.getroot().findall('.//ns:product', ns)
        length = len(products)
        counter_i = 0
        ws.cell(row=counter_i+1, column=1, value='Bundles')
        ws.cell(row=counter_i+1, column=2, value='Child Products')
        for product in products:
            counter_i += 1
            status_label.config(text="Processing Product: " + product.get('product-id'), fg="blue")
            ws.cell(row=counter_i+1, column=1, value=product.get('product-id'))
            bundled_products = product.find('./ns:bundled-products', ns)
            if bundled_products is not None:
                substring = ''
                bpl = bundled_products.findall('ns:bundled-product', ns)
                counter_j = 0
                for bundled_product in bpl:
                    counter_j += 1
                    if counter_j < len(bpl):
                        separator = ','
                    else:
                        separator = ''
                    bundled_product_id = bundled_product.attrib.get('product-id')
                    quantity = bundled_product.find('ns:quantity', ns).text
                    substring += quantity + '*' + bundled_product_id + separator
                ws.cell(row=counter_i+1, column=2, value=substring)
            percentage = ((counter_i + 1) / length) * 100
            progress_var.set(percentage)
            progress_bar.update()
            wb.save('bundlesResults.xlsx')
        return True
    except Exception as e:
        status_label_error = tk.Label(root, text="", fg="black")
        status_label_error.pack(pady=20)
        status_label_error.config(text="File processing failed! " + str(e), fg="red")
        return False


def browse_file():
    filename = filedialog.askopenfilename()
    if filename:
        progress_var.set(0)
        progress_bar.update()
        result = process_file(filename)
        if result:
            status_label.config(text="File processed successfully!", fg="green")
        else:
            status_label.config(text="File processing failed!", fg="red")


root = tk.Tk()
root.title("File Processing App")

root.geometry("750x375")

text_label = tk.Label(root, text="Select a file to process:", font=("Helvetica", 14))
text_label.pack(pady=20)

browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack()

progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate", variable=progress_var)
progress_bar.pack(pady=20)

status_label = tk.Label(root, text="", fg="black")
status_label.pack(pady=20)

root.mainloop()
